#!/usr/bin/env python3
"""Authoring tool: flatten `Brick to extended attributes.xlsx` into `brick-attributes.csv`.

The xlsx is a GPC brick -> extended-attribute matrix (one sheet per product family, one column
per brick, one row per attribute, cells marked E/R/blank). This script reconciles each attribute
row against the app's existing per-category Code List display names (the only attributes that have
CSV values + GS1 codes), and emits a flat, reviewable CSV consumed by
scripts/generate-brick-options.mjs.

Requires: openpyxl.  Run:  python3 scripts/extract-brick-attributes.py
CSV columns: category,brickCode,brickName,codeListName,requirement
"""
import csv
import os
import re
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Brick to extended attributes.xlsx")
OUT = os.path.join(ROOT, "brick-attributes.csv")

# App category -> xlsx sheet(s). Home is intentionally excluded (no xlsx coverage).
CATEGORY_SHEETS = {
    "Shoes": ["Footwear"],
    "Apparel": ["Sleepwear", "Underwear", "Swimwear"],
    "Bags": ["Handbags"],
    "Jewelry": ["Jewelry"],
    "Beauty": [
        "Cosmetics-Makeup Products", "Fragrances", "Nail Care", "Skin Care",
        "Hair Care", "Hair Removal", "Body Washing", "Sunscreen-Tanning",
    ],
}

# App display Code List Names per category (mirrors CATEGORY_ROUTING in generate-gs1-options.mjs).
# Only rows reconciling to one of these are kept, so CSV values/codes stay available downstream.
CATEGORY_DISPLAY_NAMES = {
    "Shoes": ["Shoe Type", "Shoe Style", "Closure", "Heel Type", "Heel Height Range",
              "Heel Material", "Toe Shape", "Sole Type", "Outsole Type", "Occasion",
              "Gender", "Water Repellent"],
    "Apparel": ["Code List for Dress Type", "Sleeve Type", "Collar/Neck Type", "Closure",
                "Occasion", "Gender", "Code List for Fit", "Code Type for Length Description",
                "Primary Detail Type", "Primary Detail Placement", "Primary Detail Application"],
    "Bags": ["Bag Type", "Closure", "Lining Material", "Special Embellishment",
             "Primary Detail Application", "Primary Detail Placement", "Occasion", "Gender"],
    "Jewelry": ["Jewelry Type", "Jewelry Sets", "Earring Type", "Necklace Type", "Ring Type",
                "Bracelet Type", "Band Type", "Metal", "Closure", "Occasion", "Gender"],
    "Beauty": ["Beauty Area of Use", "Beauty Treatment Specialty", "Skin Type", "Scent Type",
               "SPF Rating", "Code List for Formulation"],
}


def norm(s):
    """Collapse whitespace, lowercase, and fold the Repellant/Repellent OCR variant."""
    return re.sub(r"\s+", " ", str(s)).strip().lower().replace("repellant", "repellent")


def find_header(ws):
    """Return (codes_row, names_row). Codes row = first row (1..6) with an int in col B."""
    for r in range(1, 7):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, int):
            return r, r + 1
    raise ValueError(f"Could not locate brick code row in sheet {ws.title!r}")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    rows_out = []
    report = []

    for category, sheets in CATEGORY_SHEETS.items():
        disp_by_norm = {norm(d): d for d in CATEGORY_DISPLAY_NAMES[category]}
        cat_bricks = 0
        for sheet in sheets:
            ws = wb[sheet]
            codes_row, names_row = find_header(ws)
            # Brick columns: any col >= 2 with a non-empty name cell.
            bricks = []
            for c in range(2, ws.max_column + 1):
                name = ws.cell(row=names_row, column=c).value
                if not name or not str(name).strip():
                    continue
                code = ws.cell(row=codes_row, column=c).value
                code = str(code) if isinstance(code, int) else ""
                bricks.append((c, code, str(name).strip()))
            # Attribute rows below the names row.
            for col, code, name in bricks:
                # Skip bricks without a GPC code: they are non-image accessory/variety-pack
                # entries, and a blank code can't be a stable, unique select value downstream.
                if not code:
                    continue
                kept = []
                for r in range(names_row + 1, ws.max_row + 1):
                    a = ws.cell(row=r, column=1).value
                    if not a:
                        continue
                    disp = disp_by_norm.get(norm(a))
                    if not disp:
                        continue
                    mark = ws.cell(row=r, column=col).value
                    mark = str(mark).strip().upper() if mark else ""
                    if mark in ("E", "R"):
                        kept.append((disp, mark))
                # Require >= 2 reconciled attributes so every offered brick shows a meaningful,
                # differentiated set (drops single-"Gender" noise from thin xlsx families).
                if len(kept) < 2:
                    continue
                cat_bricks += 1
                for disp, mark in kept:
                    rows_out.append([category, code, name, disp, mark])
        report.append((category, cat_bricks))

    with open(OUT, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["category", "brickCode", "brickName", "codeListName", "requirement"])
        w.writerows(rows_out)

    print(f"Wrote {OUT} ({len(rows_out)} rows)")
    for category, n in report:
        print(f"  {category:8} -> {n} bricks with reconciled attributes")


if __name__ == "__main__":
    main()
